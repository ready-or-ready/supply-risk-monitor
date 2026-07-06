"""
페이지 3: 룰 기반 템플릿 리포트 생성 (기능4)
Section 1 자동 표시 → Section 2 입력폼 → Section 3~7 계산 결과 → 다운로드
"""
import streamlit as st
from utils.styles import apply_global_styles
import pandas as pd
import numpy as np
import io
import requests as _req
import xml.etree.ElementTree as _ET
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font

from utils.data_fetcher import (
    fetch_monthly_unit_price, fetch_country_shares,
    fetch_usd_krw_monthly, fetch_market_data, load_ncc_spread,
    fetch_news_and_notices, fetch_kotra_supply_insights, HS_CODES,
)
from utils.risk_engine import (
    calc_oil_score, calc_fx_score, calc_imp_scores,
    calc_vuln_scores, calc_geo_scores, calc_risk_results,
    top_contributors, grade_icon,
)
from utils.report_builder import (
    ncc_spread_summary, classify_scenario, build_report_text,
    calc_inventory_days, inventory_urgency,
    calc_cost_impact, delivery_price_advice,
)
from utils.llm_analyst import generate_ai_report

st.set_page_config(page_title="리포트 생성", page_icon="📄", layout="wide")
apply_global_styles()
st.title("📄 리포트 생성")
st.caption("위험지수 자동 분석 + 우리 회사 재고·원가 정보 입력 → 맞춤 리포트 출력")

# ── session_state 초기화 (페이지 이동 후에도 입력값 유지) ──
_defaults = {
    "rep_mat":            "PP",
    "rep_stock":          0.0,
    "rep_monthly_usage":  0.0,
    "rep_cost_ratio":     0.0,
    "rep_buy_price":      0.0,
    "rep_delivery_vol":   0.0,
    "rep_delivery_price": 0.0,
    "rep_submitted":      False,
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── 기간 설정 ─────────────────────────────────────────────
today      = datetime.today()
SCORE_END  = today.strftime("%Y%m")
SCORE_STRT = (today - timedelta(days=180)).strftime("%Y%m")
VULN_STRT  = (today - timedelta(days=210)).strftime("%Y%m")
VULN_END   = SCORE_END

if st.sidebar.button("🔄 데이터 새로고침"):
    st.cache_data.clear()
    st.rerun()

# ══════════════════════════════════════════════════════════
# Section 1. 이번 주 원자재 위험 요약 (자동)
# ══════════════════════════════════════════════════════════
st.subheader("Section 1. 이번 주 원자재 위험 요약")

try:
    with st.spinner("위험지수 산출 중... (관세청 API 응답에 30초 내외 소요될 수 있습니다)"):
        df_market     = fetch_market_data(days=40)
        krw_monthly   = fetch_usd_krw_monthly(SCORE_STRT, SCORE_END)
        df_ncc        = load_ncc_spread()
        df_rss, df_notice = fetch_news_and_notices()
        monthly_data  = {mat: fetch_monthly_unit_price(mat, SCORE_STRT, SCORE_END) for mat in HS_CODES}
        import_shares = {mat: fetch_country_shares(mat, VULN_STRT, VULN_END) for mat in HS_CODES}
except Exception as _e:
    st.error(
        f"데이터 수집 중 오류가 발생했습니다. 잠시 후 사이드바 **🔄 데이터 새로고침** 버튼을 눌러주세요.\n\n"
        f"_(오류: {type(_e).__name__})_"
    )
    st.stop()

oil_scores   = calc_oil_score(df_market)
fx_score     = calc_fx_score(df_market)
imp_scores   = calc_imp_scores(monthly_data, krw_monthly)
vuln_scores  = calc_vuln_scores(import_shares)
geo_scores, _, _ = calc_geo_scores(df_rss, df_notice, import_shares)
risk_results = calc_risk_results(oil_scores, fx_score, imp_scores, vuln_scores, geo_scores)
ncc_sum      = ncc_spread_summary(df_ncc)
scenario     = classify_scenario(df_market)

_grade_bg = {"낮음": "#e8f5e9", "주의": "#fffde7", "경계": "#fff3e0", "심각": "#ffebee"}
_rows_html = ""
for mat in ["PP", "PE", "PVC", "PET"]:
    r   = risk_results[mat]
    bg  = _grade_bg.get(r["등급"], "#fff")
    top = top_contributors(r)
    _rows_html += (
        f"<tr style='background:{bg}'>"
        f"<td style='padding:8px 14px;font-weight:bold'>{mat}</td>"
        f"<td style='padding:8px 14px;text-align:center;font-size:20px;font-weight:bold'>{r['위험지수']:.1f}</td>"
        f"<td style='padding:8px 14px;text-align:center'>{grade_icon(r['등급'])} {r['등급']}</td>"
        f"<td style='padding:8px 14px;font-size:12px;color:#555'>{' / '.join(top[:2])}</td>"
        f"</tr>"
    )
st.markdown(
    f"<table style='width:100%;border-collapse:collapse;border:1px solid #e0e0e0'>"
    f"<thead><tr style='background:#f5f5f5;border-bottom:2px solid #ddd'>"
    f"<th style='padding:8px 14px;text-align:left'>품목</th>"
    f"<th style='padding:8px 14px;text-align:center'>위험지수</th>"
    f"<th style='padding:8px 14px;text-align:center'>등급</th>"
    f"<th style='padding:8px 14px;text-align:left'>주요 상승 요인</th>"
    f"</tr></thead><tbody>{_rows_html}</tbody></table>",
    unsafe_allow_html=True,
)
st.caption(f"기준: {today.strftime('%Y-%m-%d')} | 등급 기준: 2020–2024년 과거 분포")

st.info(f"🔍 나프타 시나리오: **{scenario['scenario_name']}** — {scenario['scenario_desc']}")
st.caption(f"역사적 NCC 스프레드 구간: {ncc_sum['스프레드']} USD/ton ({ncc_sum['기준분기']}) — {ncc_sum['구간_텍스트']}")

# KOTRA 글로벌 공급망 인사이트
gsi_items = fetch_kotra_supply_insights(num=2)
if gsi_items:
    st.markdown("**📦 KOTRA 글로벌 공급망 인사이트 (최신)**")
    for item in gsi_items:
        cols_k = st.columns([6, 2, 2])
        cols_k[0].markdown(f"[{item['제목']}]({item['링크']})")
        cols_k[1].caption(item["날짜"])
        if item["pdf"]:
            cols_k[2].markdown(f"[PDF]({item['pdf']})")

st.markdown("---")

# ══════════════════════════════════════════════════════════
# Section 2. 우리 회사 영향 예상 — 입력폼
# ══════════════════════════════════════════════════════════
st.subheader("Section 2. 우리 회사 영향 예상 입력")
st.caption("필수 항목만 입력해도 Section 3(재고소진) 결과를 볼 수 있습니다. 입력값은 다른 탭을 갔다 와도 유지됩니다.")

mat_options  = list(HS_CODES.keys())
selected_mat = st.selectbox(
    "분석할 원재료",
    mat_options,
    index=mat_options.index(st.session_state["rep_mat"]),
    key="rep_mat",
)

with st.form("input_form"):
    st.markdown("**필수 입력**")
    col1, col2 = st.columns(2)
    with col1:
        stock = st.number_input(
            "현재 재고량 (톤)", min_value=0.0, step=0.1,
            value=st.session_state["rep_stock"],
            help="현재 보유 중인 원재료 재고량",
        )
    with col2:
        monthly_usage = st.number_input(
            "월 원재료 사용량 (톤)", min_value=0.0, step=0.1,
            value=st.session_state["rep_monthly_usage"],
            help="한 달에 소비하는 원재료 양",
        )

    st.markdown("**선택 입력 — 원가 분석**")
    col3, col4 = st.columns(2)
    with col3:
        cost_ratio = st.number_input(
            "원재료별 원가 비중 (%)", min_value=0.0, max_value=100.0, step=0.1,
            value=st.session_state["rep_cost_ratio"],
            help="완제품 제조원가에서 해당 원재료가 차지하는 비율",
        )
    with col4:
        recent_buy_price = st.number_input(
            "최근 매입단가 (원/kg)", min_value=0.0, step=1.0,
            value=st.session_state["rep_buy_price"],
            help="가장 최근 실제 매입 단가. 미입력 시 관세청 기준 수입단가로 대체",
        )

    st.markdown("**선택 입력 — 납품단가 분석**")
    col5, col6 = st.columns(2)
    with col5:
        delivery_volume = st.number_input(
            "월 납품 물량 (톤)", min_value=0.0, step=0.1,
            value=st.session_state["rep_delivery_vol"],
        )
    with col6:
        delivery_price = st.number_input(
            "평균 납품단가 (원/개 또는 원/kg)", min_value=0.0, step=1.0,
            value=st.session_state["rep_delivery_price"],
            help="완제품 납품 시 받는 단가. 미입력 시 방향만 표시",
        )

    submitted = st.form_submit_button("📊 리포트 생성", type="primary")

# 폼 제출 시 session_state 저장
if submitted:
    st.session_state["rep_stock"]          = stock
    st.session_state["rep_monthly_usage"]  = monthly_usage
    st.session_state["rep_cost_ratio"]     = cost_ratio
    st.session_state["rep_buy_price"]      = recent_buy_price
    st.session_state["rep_delivery_vol"]   = delivery_volume
    st.session_state["rep_delivery_price"] = delivery_price
    st.session_state["rep_submitted"]      = True

# 제출된 적 없으면 아래 결과 섹션 미표시
if not st.session_state["rep_submitted"]:
    st.stop()

# 결과 섹션에서 사용할 값은 session_state에서 읽음
_stock          = st.session_state["rep_stock"]
_monthly_usage  = st.session_state["rep_monthly_usage"]
_cost_ratio     = st.session_state["rep_cost_ratio"]
_buy_price      = st.session_state["rep_buy_price"]
_delivery_price = st.session_state["rep_delivery_price"]
_mat            = st.session_state["rep_mat"]

# ══════════════════════════════════════════════════════════
# Section 3. 재고 소진 예상
# ══════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Section 3. 재고 소진 예상")

inv_days = calc_inventory_days(_stock, _monthly_usage) if _stock > 0 and _monthly_usage > 0 else None
urgency_text, urgency_icon = inventory_urgency(inv_days)

col_a, col_b = st.columns(2)
with col_a:
    if inv_days is not None:
        st.metric("재고 소진 예상일", f"{inv_days:.0f}일")
        st.caption(f"재고 소진 예상일 = 현재 재고량 ÷ 일평균 사용량  ({_stock:,.0f}kg ÷ {_monthly_usage/30:,.1f}kg/일)")
    else:
        st.warning("재고량 또는 월 사용량이 입력되지 않았습니다.")
with col_b:
    st.markdown(f"### {urgency_icon} {urgency_text}")
    if inv_days is not None and inv_days < 30:
        st.error(f"현재 재고 여유가 {inv_days:.0f}일분입니다. 조달 일정을 확인하세요.")

st.caption(
    "**긴급도 기준**　🔴 7일 미만: 즉시 발주 검토　"
    "🟠 8\~14일: 발주 일정 확인　"
    "🟡 15\~30일: 모니터링　"
    "🟢 30일 초과: 여유"
)

# ══════════════════════════════════════════════════════════
# Section 4. 원가 상승 추정
# ══════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Section 4. 원가 상승 추정")

df_m = monthly_data[_mat].copy()
df_m["환율_KRW"]        = df_m["연월"].map(krw_monthly)
df_m["수입단가_KRW_kg"] = df_m["수입단가_USD_kg"] * df_m["환율_KRW"]
df_m["ln"]              = np.log(df_m["수입단가_KRW_kg"]).diff()
valid = df_m[df_m["ln"].notna()]

imp_change = None
gap        = None

col_ref, col_gap = st.columns(2)

if not valid.empty:
    imp_change       = (np.exp(valid.iloc[-1]["ln"]) - 1) * 100
    latest_mon       = valid.iloc[-1]["연월"]
    market_price_krw = valid.iloc[-1]["수입단가_KRW_kg"]

    with col_ref:
        st.metric(
            f"관세청 기준 수입단가 ({latest_mon})",
            f"{market_price_krw:,.0f} 원/kg",
            delta=f"전월 대비 {imp_change:+.2f}%",
            delta_color="inverse" if imp_change > 0 else "normal",
        )
        st.caption("관세청 후행 특성상 1\~2개월 전 실거래 확정 데이터")

    if _buy_price > 0 and market_price_krw > 0:
        gap = (_buy_price - market_price_krw) / market_price_krw * 100
        with col_gap:
            st.metric(
                "우리 회사 매입단가 vs 관세청 기준 괴리",
                f"{gap:+.2f}%",
            )
            sign = "비싸게" if gap > 0 else "저렴하게"
            st.caption(
                f"우리 매입가 {_buy_price:,.0f}원/kg  vs  관세청 기준 {market_price_krw:,.0f}원/kg  "
                f"→ 시장가보다 **{abs(gap):.1f}%** {sign}매입 중"
            )
    else:
        with col_gap:
            st.info("최근 매입단가를 입력하면 시장가 대비 괴리율을 확인할 수 있습니다.")
else:
    st.warning("수입단가 데이터 부족")

# 예상 원가 영향 — 항상 관세청 전월비 기준
cost_impact = None
if _cost_ratio > 0 and imp_change is not None:
    cost_impact = calc_cost_impact(imp_change, _cost_ratio)
    st.metric("예상 원가 영향", f"{cost_impact:+.2f}%")
    st.caption(f"= 관세청 수입단가 전월 변동 ({imp_change:+.2f}%) × 원가 비중 ({_cost_ratio:.1f}%)")
elif _cost_ratio == 0:
    st.info("원재료별 원가 비중을 입력하면 원가 영향을 산출합니다.")

# 현재 조달 부담 — 매입단가 입력 시에만 표시
if gap is not None and _cost_ratio > 0:
    procurement_burden = calc_cost_impact(gap, _cost_ratio)
    burden_icon = "🔴" if procurement_burden > 2 else ("🟡" if procurement_burden > 0 else "🟢")
    st.metric(
        "현재 조달 부담",
        f"{procurement_burden:+.2f}%",
        help="우리 매입단가 기준 — 시장가보다 비싸면 양수(원가 부담), 싸면 음수(원가 절감)",
    )
    st.caption(
        f"{burden_icon} = 매입단가 괴리 ({gap:+.2f}%) × 원가 비중 ({_cost_ratio:.1f}%)  "
        "— 관세청 기준가로 샀을 때 대비 원가율 차이"
    )

# ══════════════════════════════════════════════════════════
# Section 5. 납품단가 조정 필요성
# ══════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Section 5. 납품단가 조정 필요성")

advice, reason = delivery_price_advice(cost_impact)

col5a, col5b = st.columns([1, 1])

with col5a:
    st.markdown(f"### 💬 {advice}")
    if reason:
        st.caption(reason)
    if _delivery_price > 0 and cost_impact is not None:
        estimated_new = _delivery_price * (1 + cost_impact / 100)
        st.metric("납품단가 조정 시 예상 단가", f"{estimated_new:,.0f}원",
                  delta=f"+{_delivery_price * cost_impact / 100:,.0f}원 ({cost_impact:+.2f}%)")
    elif _delivery_price == 0:
        st.caption("조정 예상 단가 산출: 원가 비중 + 평균 납품단가 둘 다 입력 시 산출됩니다.")

with col5b:
    st.markdown("**판단 기준 (예상 원가 영향 기준)**")
    rules = [
        ("7% 이상",  "단가 조정 근거자료 준비 권고", "중소 제조업 평균 영업이익률 5\~8% 실질 잠식 구간"),
        ("3\~7%",    "납품단가 협의 검토",           "하도급법 제16조·상생협력법 제22조의2 협의 신청 구간"),
        ("0\~3%",    "모니터링 유지",                "하도급법 제16조 협의 신청 기준 미달"),
        ("0% 미만",  "원가 개선 구간",               "수입단가 하락 — 원가율 완화 예상"),
    ]
    for threshold, label, desc in rules:
        is_current = (label == advice)
        prefix = "▶ " if is_current else "　"
        weight = "**" if is_current else ""
        st.markdown(
            f"{prefix}{weight}{threshold} → {label}{weight}  \n"
            f"<span style='color:gray;font-size:0.82em'>{desc}</span>",
            unsafe_allow_html=True,
        )

# ══════════════════════════════════════════════════════════
# Section 6. 정책지원 공고
# ══════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Section 6. 관련 정책 지원")
st.caption("중기부 공고 기준 · 최근 60일 이내 공고 · 키워드: 원재료 · 조달 · 구매 · 수입 · 공급망")

_SME_KEYWORDS = ["원재료", "조달", "구매", "수입", "공급망"]

try:
    sme_key = st.secrets.get("SME_KEY", "")
    if not sme_key:
        st.info("중기부 API 키(SME_KEY)가 설정되지 않았습니다.")
    else:
        from datetime import datetime as _dt, timedelta as _td

        res = _req.get(
            "https://apis.data.go.kr/1421000/bizinfo/pblancBsnsService",
            params={"serviceKey": sme_key, "numOfRows": 100, "pageNo": 1},
            timeout=15,
        )
        root      = _ET.fromstring(res.text)
        all_items = root.findall(".//item")
        cutoff    = _dt.today() - _td(days=60)

        def _created_within_60d(it):
            raw = (it.findtext("creatPnttm") or "")[:10]
            try:
                return _dt.strptime(raw, "%Y-%m-%d") >= cutoff
            except Exception:
                return True  # 날짜 파싱 실패 시 포함

        recent = [it for it in all_items if _created_within_60d(it)]
        filtered = [
            it for it in recent
            if any(kw in (it.findtext("pblancNm") or "") for kw in _SME_KEYWORDS)
        ]

        # 키워드 매칭 없으면 60일 이내 최신 5건, 그것도 없으면 전체 최신 5건
        if filtered:
            show_items   = filtered[:5]
            fallback_msg = None
        elif recent:
            show_items   = recent[:5]
            fallback_msg = "최근 60일 이내 관련 키워드 공고가 없어 최신 공고를 표시합니다."
        else:
            show_items   = all_items[:5]
            fallback_msg = "최근 60일 이내 공고가 없어 전체 최신 공고를 표시합니다."

        if fallback_msg:
            st.caption(fallback_msg)

        for item in show_items:
            name   = item.findtext("pblancNm") or "제목 없음"
            period = item.findtext("reqstBeginEndDe") or "기간 정보 없음"
            inst   = item.findtext("jrsdInsttNm") or ""
            target = item.findtext("trgetNm") or ""
            url    = item.findtext("pblancUrl") or ""
            realm  = item.findtext("pldirSportRealmLclasCodeNm") or ""

            with st.expander(f"📌 {name}"):
                col_i, col_ii = st.columns(2)
                col_i.markdown(f"**소관기관:** {inst}")
                col_ii.markdown(f"**지원분야:** {realm}")
                st.markdown(f"**신청기간:** {period}  |  **지원대상:** {target}")
                if url:
                    st.markdown(f"[🔗 공고 바로가기]({url})")

except Exception as e:
    st.info(f"정책지원 공고 조회 중 오류: {e}")

# ══════════════════════════════════════════════════════════
# Section 7. 협상 방향 신호카드
# ══════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Section 7. 협상 방향 신호카드")
st.caption("협상시 접근할 방향에 대한 참고로 활용해보세요.")

r          = risk_results[_mat]
grade      = r["등급"]
imp_ch_str = f"{imp_change:+.2f}%" if imp_change is not None else "N/A"

# 외부 시장 데이터 계산
_brent = df_market["Brent"].dropna() if "Brent" in df_market.columns else pd.Series(dtype=float)
brent_20d_ret = (_brent.iloc[-1] / _brent.iloc[-21] - 1) * 100 if len(_brent) >= 21 else None
_fx = df_market["USDKRW"].dropna() if "USDKRW" in df_market.columns else pd.Series(dtype=float)
usdkrw_20d_chg = (_fx.iloc[-1] / _fx.iloc[-21] - 1) * 100 if len(_fx) >= 21 else None
usdkrw_recent  = _fx.iloc[-1] if len(_fx) >= 1 else None

# 신호 방향 결정 (내부 위험지수 등급 기준)
_b_str = f" / 유가 20일 {brent_20d_ret:+.1f}%" if brent_20d_ret is not None else ""
if grade in ("경계", "심각") and imp_change is not None and imp_change > 0:
    signal = "📈 납품단가 인상 요청 타이밍"
    signal_detail = (
        f"관세청 수입단가 {imp_ch_str} 상승{_b_str} — 원가 압력이 가중되는 구간입니다. "
        "시장 데이터를 근거자료로 준비해 선제 제시하세요."
    )
elif grade in ("경계", "심각"):
    signal = "🔒 장기계약 · 물량 확보 협상 제안"
    signal_detail = (
        f"수입단가 {imp_ch_str}{_b_str} — 가격보다 공급 안정성이 우선인 구간입니다. "
        "물량 확보 조건 중심으로 협상하세요."
    )
elif grade in ("낮음", "주의") and imp_change is not None and imp_change > 0:
    signal = "🛡️ 공급처 인상 요구 방어"
    signal_detail = (
        f"관세청 수입단가 {imp_ch_str} 상승이 있으나 공급 리스크는 낮은 구간{_b_str}. "
        "시장 데이터로 인상 폭을 제한하는 협상을 진행하세요."
    )
else:
    signal = "📉 현 조건 유지 또는 단가 인하 협상 검토"
    signal_detail = (
        f"수입단가 {imp_ch_str}{_b_str} — 원가 압력이 낮은 구간으로 현 조건이 유리합니다."
    )

st.markdown(f"### ▶ {signal}")
st.write(signal_detail)

col_int, col_ext = st.columns(2)

with col_int:
    st.markdown("**내부 판단 근거 — 위험지수 세부점수**")
    st.caption("신호 방향 결정에 사용한 내부 지표 (협상 테이블에는 제시하지 않습니다)")
    st.metric("종합 위험지수", f"{r['위험지수']:.1f}점", grade)
    for label, key, weight in [
        ("국제유가 압력",   "유가",     "25%"),
        ("환율 압력",       "환율",     "15%"),
        ("수입가격 압력",   "수입가격", "20%"),
        ("수입구조 취약성", "수입구조", "20%"),
        ("지정학 노출",     "지정학",   "20%"),
    ]:
        score   = r.get(key, 0)
        bar_pct = min(int(score), 100)
        color   = "#ef5350" if score >= 60 else ("#ff9800" if score >= 35 else "#66bb6a")
        st.markdown(
            f"<div style='margin:3px 0'>"
            f"<span style='font-size:0.84em'>{label} <span style='color:#aaa'>({weight})</span></span>"
            f"<div style='background:#eee;border-radius:4px;height:7px;margin:2px 0'>"
            f"<div style='background:{color};width:{bar_pct}%;height:7px;border-radius:4px'></div></div>"
            f"<span style='font-size:0.8em;color:#555'>{score:.1f}점</span>"
            f"</div>",
            unsafe_allow_html=True,
        )

with col_ext:
    st.markdown("**협상 시 제시 가능한 외부 데이터**")
    st.caption("공급처와의 협상에서 외부 공인 데이터로 활용 가능")
    ext_rows = [("관세청 수입단가 전월비", imp_ch_str, "관세청")]
    if brent_20d_ret is not None:
        ext_rows.append(("Brent 유가 20일 변동", f"{brent_20d_ret:+.1f}%", "yfinance(BZ=F)"))
    if usdkrw_recent is not None:
        fx_str = f"{usdkrw_recent:,.0f}원"
        if usdkrw_20d_chg is not None:
            fx_str += f"  ({usdkrw_20d_chg:+.1f}% 20일)"
        ext_rows.append(("원/달러 환율", fx_str, "yfinance(USDKRW=X)"))
    if cost_impact is not None:
        ext_rows.append(("예상 원가 영향", f"{cost_impact:+.2f}%", "관세청 + 입력 데이터"))
    if inv_days is not None:
        ext_rows.append(("재고 여유", f"{inv_days:.0f}일  {urgency_icon}", "입력 데이터"))
    for label, val, source in ext_rows:
        st.markdown(
            f"<div style='margin:5px 0;padding:7px 10px;background:#f8f9fa;border-radius:6px'>"
            f"<span style='font-size:0.82em;color:#666'>{label}</span><br>"
            f"<span style='font-weight:bold'>{val}</span>"
            f"<span style='font-size:0.75em;color:#bbb;float:right'>출처: {source}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )

st.caption("※ 협상 구체적 요청 폭(버퍼 비율)은 추후 구매 협상 전문가 자문 후 고도화 예정")

# ══════════════════════════════════════════════════════════
# Section 8. AI 종합 분석
# ══════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("Section 8. AI 종합 분석")
st.caption("위험지수·시장 데이터·귀사 현황을 종합한 자연어 분석을 생성합니다.")

if st.button("🤖 AI 종합 분석 생성", type="primary"):
    with st.spinner("분석 중..."):
        _ai_text, _is_live = generate_ai_report(
            mat=_mat,
            r=risk_results[_mat],
            ncc_sum=ncc_sum,
            scenario=scenario,
            inv_days=inv_days,
            cost_impact=cost_impact,
            imp_change_str=imp_ch_str,
            advice=advice,
            df_rss=df_rss,
        )
    st.session_state["ai_report_text"] = _ai_text
    st.session_state["ai_report_live"] = _is_live

if "ai_report_text" in st.session_state:
    if not st.session_state.get("ai_report_live", True):
        st.caption("ℹ️ OPENAI_KEY 미설정 — 샘플 분석 표시 중")
    st.markdown(st.session_state["ai_report_text"])

# ══════════════════════════════════════════════════════════
# 다운로드
# ══════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("⬇️ 리포트 다운로드")

user_inputs = {}
if inv_days is not None:
    user_inputs["재고_소진_예상일"] = inv_days
if cost_impact is not None:
    user_inputs["원가_영향_%"] = cost_impact

report_text = build_report_text(_mat, r, ncc_sum, scenario, user_inputs or None)
today_str   = datetime.today().strftime("%Y%m%d")
today_label = datetime.today().strftime("%Y-%m-%d")
ai_text     = st.session_state.get("ai_report_text", "")

# ── TXT: AI 분석 append ───────────────────────────────────
txt_data = report_text
if ai_text:
    txt_data += "\n\n" + "="*60 + "\n■ AI 종합 분석\n" + "="*60 + "\n" + ai_text

# ── Excel ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"{_mat} 지표"
ws.append([f"[{_mat}] 원자재 리스크 모니터링 리포트", today_label])
ws["A1"].font = Font(bold=True, size=13)
ws.append([])

_signal_clean = signal.replace("📈","").replace("🔒","").replace("🛡️","").replace("📉","").strip()
for row in [
    ("위험지수",       r["위험지수"],               ""),
    ("등급",           r["등급"],                   ""),
    ("국제유가 압력",  r["유가"],                   "점 (가중치 25%)"),
    ("환율 압력",      r["환율"],                   "점 (가중치 15%)"),
    ("수입가격 압력",  r["수입가격"],               "점 (가중치 20%)"),
    ("수입구조 취약성", r["수입구조"],              "점 (가중치 20%)"),
    ("지정학 노출",    r["지정학"],                 "점 (가중치 20%)"),
    ("", "", ""),
    ("나프타 시나리오", scenario.get("scenario_name", ""), scenario.get("scenario_desc", "")),
    ("NCC 스프레드",   ncc_sum["스프레드"],         f"USD/ton — {ncc_sum['구간_텍스트']}"),
    ("", "", ""),
    ("재고 소진 예상일", inv_days if inv_days else "미입력", "일"),
    ("원가 영향",      f"{cost_impact:+.2f}%" if cost_impact else "미입력", ""),
    ("납품단가 권고",  advice,                      ""),
    ("협상 방향 신호", _signal_clean,               ""),
]:
    ws.append(list(row))

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 55

if ai_text:
    ws_ai = wb.create_sheet("AI 종합분석")
    ws_ai.append([f"[{_mat}] AI 종합 분석", today_label])
    ws_ai["A1"].font = Font(bold=True, size=13)
    ws_ai.append([])
    for line in ai_text.splitlines():
        clean = line.replace("**", "").replace("①", "1.").replace("②", "2.").replace("③", "3.")
        ws_ai.append([clean])
    ws_ai.column_dimensions["A"].width = 100

buf = io.BytesIO()
wb.save(buf)
buf.seek(0)

# ── HTML ──────────────────────────────────────────────────
def _row(label, val, note=""):
    note_td = f"<td style='color:#777;font-size:0.9em'>{note}</td>" if note else "<td></td>"
    return f"<tr><td style='font-weight:bold;width:180px'>{label}</td><td>{val}</td>{note_td}</tr>"

ai_section = ""
if ai_text:
    ai_html = ai_text.replace("\n", "<br>").replace("**", "").replace(
        "①", "<b>①").replace("②", "<b>②").replace("③", "<b>③")
    ai_section = f"""
    <h2>AI 종합 분석</h2>
    <div style='line-height:1.8'>{ai_html}</div>"""

html_data = f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <title>[{_mat}] 원자재 리스크 리포트 {today_label}</title>
  <style>
    body {{ font-family: 'Malgun Gothic', '맑은 고딕', sans-serif; max-width: 860px;
            margin: 0 auto; padding: 30px; color: #222; }}
    h1   {{ font-size: 1.4em; border-bottom: 2px solid #333; padding-bottom: 8px; }}
    h2   {{ font-size: 1.1em; margin-top: 28px; border-left: 4px solid #1976d2;
            padding-left: 10px; color: #1976d2; }}
    table {{ border-collapse: collapse; width: 100%; margin: 8px 0; }}
    td, th {{ border: 1px solid #ddd; padding: 7px 10px; vertical-align: top; }}
    tr:nth-child(even) {{ background: #f9f9f9; }}
    .meta {{ color: #777; font-size: 0.88em; margin-bottom: 20px; }}
    @media print {{ body {{ padding: 10px; }} }}
  </style>
</head>
<body>
  <h1>[{_mat}] 원자재 리스크 모니터링 리포트</h1>
  <p class="meta">기준일: {today_label}</p>

  <h2>위험지수</h2>
  <table>
    {_row("종합 위험지수", f"{r['위험지수']:.1f}점", r['등급'])}
    {_row("국제유가 압력", f"{r['유가']:.1f}점", "가중치 25%")}
    {_row("환율 압력", f"{r['환율']:.1f}점", "가중치 15%")}
    {_row("수입가격 압력", f"{r['수입가격']:.1f}점", "가중치 20%")}
    {_row("수입구조 취약성", f"{r['수입구조']:.1f}점", "가중치 20%")}
    {_row("지정학 노출", f"{r['지정학']:.1f}점", "가중치 20%")}
  </table>

  <h2>나프타 시장</h2>
  <table>
    {_row("시나리오", scenario.get('scenario_name',''), scenario.get('scenario_desc',''))}
    {_row("NCC 스프레드", f"{ncc_sum['스프레드']} USD/ton", ncc_sum['구간_텍스트'])}
  </table>

  <h2>귀사 현황</h2>
  <table>
    {_row("재고 소진 예상일", f"{inv_days:.0f}일" if inv_days else "미입력")}
    {_row("예상 원가 영향", f"{cost_impact:+.2f}%" if cost_impact else "미입력")}
    {_row("납품단가 권고", advice)}
    {_row("협상 방향 신호", _signal_clean)}
  </table>
  {ai_section}
</body>
</html>"""

col_dl1, col_dl2, col_dl3 = st.columns(3)

with col_dl1:
    st.download_button(
        label="📝 TXT 다운로드",
        data=txt_data.encode("utf-8"),
        file_name=f"report_{_mat}_{today_str}.txt",
        mime="text/plain",
    )

with col_dl2:
    st.download_button(
        label="📊 Excel 다운로드",
        data=buf.getvalue(),
        file_name=f"report_{_mat}_{today_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

with col_dl3:
    st.download_button(
        label="🌐 HTML 다운로드 (인쇄·PDF용)",
        data=html_data.encode("utf-8"),
        file_name=f"report_{_mat}_{today_str}.html",
        mime="text/html",
    )
    st.caption("브라우저에서 열고 Ctrl+P → PDF로 저장")
